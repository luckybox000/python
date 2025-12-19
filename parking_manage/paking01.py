import os
from datetime import datetime, timedelta
from openpyxl import Workbook

# 1. 파일들을 저장할 폴더 생성
folder_name = "일자별_로그_보고서"
if not os.path.exists(folder_name):
    os.makedirs(folder_name)
    print(f"'{folder_name}' 폴더를 생성했습니다.")

# 2. 시작 날짜 설정 (오늘 기준)
start_date = datetime.now()

# 3. 100일 동안 반복하며 파일 생성
for i in range(100):
    # 시작 날짜에 i일을 더함
    current_date = start_date + timedelta(days=i)
    
    # 날짜를 문자열로 변환 (예: 2023-10-27)
    date_str = current_date.strftime('%Y-%m-%d')
    
    # 새 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "일일보고서"
    
    # 파일 내부에 날짜 정보 기입 (선택 사항)
    ws['A1'] = "작성일자"
    ws['B1'] = date_str
    
    # 4. 파일 저장 (예: 2023-10-27.xlsx)
    file_name = f"{date_str}.xlsx"
    file_path = os.path.join(folder_name, file_name)
    
    wb.save(file_path)

print(f"오늘부터 100일치({date_str}까지) 파일 생성이 모두 완료되었습니다!")
