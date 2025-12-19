import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 1. 워크북 및 시트 설정
wb = Workbook()
ws = wb.active
ws.title = "자동계산_주차관리"

# 2. 헤더 설정
headers = ["순번", "차량번호", "입차시간", "출차시간", "주차시간(분)", "주차요금(원)", "비고"]
ws.append(headers)

# 요금 규정 설정 (예: 10분당 500원)
price_per_10min = 500

# 3. 데이터 및 수식 입력 (10개 행 미리 만들기)
# 엑셀 수식 설명:
# 주차시간: (출차-입차) * 1440 (하루의 분 단위 환산)
# 주차요금: (주차시간 / 10) * 500원
for i in range(2, 12):  # 2행부터 11행까지
    row_num = str(i)
    ws[f'A{row_num}'] = i - 1  # 순번
    
    # 주차시간 계산 수식 (E열) : 출차(D) - 입차(C)
    # 엑셀에서 날짜 빼기는 '일' 단위이므로 1440을 곱해 '분'으로 변환
    ws[f'E{row_num}'] = f'=IF(AND(C{row_num}<>"", D{row_num}<>""), ROUND((D{row_num}-C{row_num})*1440, 0), 0)'
    
    # 주차요금 계산 수식 (F열) : (주차시간/10) * 요금
    ws[f'F{row_num}'] = f'=IF(E{row_num}>0, ROUNDUP(E{row_num}/10, 0)*{price_per_10min}, 0)'

# 4. 스타일 적용 (가독성)
header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid") # 연녹색
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 열 너비 조정
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15

# 5. 파일 저장
file_name = "스마트_주차관리_시스템.xlsx"
wb.save(file_name)

print(f"'{file_name}' 파일이 생성되었습니다. 입/출차 시간 입력 시 자동 계산됩니다.")