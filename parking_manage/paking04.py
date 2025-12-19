import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference

# 1. 워크북 생성 및 기본 시트 설정 (Sheet1: 상세 내역)
wb = Workbook()
ws1 = wb.active
ws1.title = "주차상세내역"

# [Sheet1] 헤더 및 자동 수식 설정 (이전과 동일)
headers = ["순번", "차량번호", "입차시간", "출차시간", "주차시간(분)", "주차요금(원)", "비고"]
ws1.append(headers)

price_per_10min = 500
for i in range(2, 52):  # 50행까지 생성
    row_num = str(i)
    ws1[f'A{row_num}'] = i - 1
    ws1[f'E{row_num}'] = f'=IF(AND(C{row_num}<>"", D{row_num}<>""), ROUND((D{row_num}-C{row_num})*1440, 0), 0)'
    ws1[f'F{row_num}'] = f'=IF(E{row_num}>0, ROUNDUP(E{row_num}/10, 0)*{price_per_10min}, 0)'

# 2. 통계 시트 생성 (Sheet2: 월별 통계)
ws2 = wb.create_sheet("월별통계")

# [Sheet2] 통계 테이블 헤더
stats_headers = ["월", "총 주차 대수", "총 수익(원)"]
ws2.append(stats_headers)

# 샘플 월 데이터 입력 (1월 ~ 12월)
months = [f"{m}월" for m in range(1, 13)]
for m_text in months:
    ws2.append([m_text, 0, 0]) # 초기값 0

# [Sheet2] 스타일 적용
header_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid") # 연파랑
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# 3. 차트(그래프) 추가
# 주차 수익 통계를 보여주는 막대 그래프 생성
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "월별 주차 수익 현황"
chart.y_axis.title = "수익(원)"
chart.x_axis.title = "월"

# 데이터 범위 설정 (수익 데이터: C2 ~ C13)
data = Reference(ws2, min_col=3, min_row=1, max_row=13)
# 범주 설정 (월 이름: A2 ~ A13)
cats = Reference(ws2, min_col=1, min_row=2, max_row=13)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4

# 그래프를 Sheet2의 E2 위치에 배치
ws2.add_chart(chart, "E2")

# 4. 마무리: 열 너비 및 저장
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 20
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15

file_name = "스마트_주차관리_v2.xlsx"
wb.save(file_name)

print(f"'{file_name}'이 생성되었습니다. Sheet2에서 통계 그래프를 확인하세요!")